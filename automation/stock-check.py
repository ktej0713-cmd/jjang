# -*- coding: utf-8 -*-
"""
짱베이스볼 야구바지 재고 자동 모니터링
Windows 예약 작업(schtasks)으로 매일 오전 8시 실행

동작:
  1. 고도몰 Open API로 24개 상품 옵션별 재고 조회
  2. 엑셀 파일에서 주문중/60일 필요량 읽기
  3. 엑셀 현재 재고 열을 API 데이터로 업데이트
  4. 추가 발주 자동 계산
  5. HTML 리포트 생성 + Gmail 발송

사용법:
  python stock-check.py          # 전체 조회 + 엑셀 업데이트 + 메일 발송
  python stock-check.py --nomail # 메일 없이 엑셀만 업데이트
  python stock-check.py --test   # 테스트 (1개 상품만 조회)
"""

import json
import sys
import logging
import smtplib
import shutil
import xml.etree.ElementTree as ET
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

try:
    import requests
except ImportError:
    print("[ERROR] requests 패키지 필요: pip install requests")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("[ERROR] openpyxl 패키지 필요: pip install openpyxl")
    sys.exit(1)

# === 경로 설정 ===
SCRIPT_DIR = Path(__file__).parent
CONFIG_PATH = SCRIPT_DIR.parent / "config" / "stock-monitor.json"
OUTPUT_DIR = SCRIPT_DIR.parent / "output"
LOG_DIR = SCRIPT_DIR / "logs"

# 엑셀 파일 경로 (원본)
EXCEL_PATH = Path.home() / "짱베이스볼_야구바지_재고발주_v5_20260319.xlsx"
BACKUP_DIR = Path.home() / "backup"

# 고도몰5 Open API 엔드포인트
GODOMALL_API_URL = "https://openhub.godo.co.kr/godomall5/goods/Goods_Search.php"

# 사이즈 열 매핑 (엑셀 기준)
SIZES = ["28", "30", "32", "34", "36", "38", "40"]
# 현재 재고: C4(합계), C5~C11(28~40)
STOCK_COL_START = 5    # C5 = 28
STOCK_TOTAL_COL = 4    # C4 = 합계
# 주문중: C12(합계), C13~C19(28~40)
ORDER_COL_START = 13   # C13 = 28
ORDER_TOTAL_COL = 12   # C12 = 합계
# 60일 필요량: C20(합계), C21~C27(28~40)
NEED_COL_START = 21    # C21 = 28
NEED_TOTAL_COL = 20    # C20 = 합계
# 추가 발주: C28(합계), C29~C35(28~40)
REORDER_COL_START = 29 # C29 = 28
REORDER_TOTAL_COL = 28 # C28 = 합계

# 디렉토리 생성
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

# === 로깅 설정 ===
log_file = LOG_DIR / f"stock-check-{datetime.now().strftime('%Y%m%d')}.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)


def load_config():
    """설정 파일 로드"""
    if not CONFIG_PATH.exists():
        log.error(f"설정 파일 없음: {CONFIG_PATH}")
        sys.exit(1)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = json.load(f)
    partner_key = config.get("godomall", {}).get("partner_key", "")
    if not partner_key or partner_key == "여기에_파트너키_입력":
        log.error("고도몰 partner_key가 설정되지 않았습니다")
        sys.exit(1)
    return config


def fetch_stock_api(config):
    """고도몰5 Open API로 상품별 옵션/재고 조회"""
    partner_key = config["godomall"]["partner_key"]
    secret_key = config["godomall"]["secret_key"]
    products = config["monitor_products"]

    # 고유 goodsNo만 추출 (중복 제거)
    unique_goods = {}
    for prod in products:
        gno = prod["goodsNo"]
        if gno not in unique_goods:
            unique_goods[gno] = prod

    results = {}
    session = requests.Session()

    for gno, prod in unique_goods.items():
        log.info(f"조회중: {prod['name']} (goodsNo: {gno})")
        try:
            resp = session.post(
                GODOMALL_API_URL,
                data={"partner_key": partner_key, "key": secret_key, "goodsNo": gno},
                timeout=15
            )
            if resp.status_code != 200:
                log.error(f"  -> API HTTP {resp.status_code}")
                continue

            root = ET.fromstring(resp.text)
            code = root.findtext(".//header/code", "")
            if code != "000":
                log.error(f"  -> API 에러: {root.findtext('.//header/msg', '')}")
                continue

            goods = root.find(".//goods_data")
            if goods is None:
                continue

            # 옵션별 재고 딕셔너리
            sizes = {}
            for opt in goods.findall("optionData"):
                opt_val = opt.findtext("optionValue1", "")
                stock_cnt = int(opt.findtext("stockCnt", "0"))
                if opt_val:
                    sizes[opt_val] = stock_cnt

            total = sum(sizes.values())
            results[gno] = {"sizes": sizes, "total": total}
            log.info(f"  -> API 성공: {len(sizes)}개 옵션, 총 재고 {total}")

        except Exception as e:
            log.error(f"  -> API 요청 실패: {e}")

    return results


def update_excel(api_data):
    """엑셀 파일의 현재 재고를 API 데이터로 업데이트하고 분석 결과 반환"""

    if not EXCEL_PATH.exists():
        log.error(f"엑셀 파일 없음: {EXCEL_PATH}")
        return None

    # 백업
    backup_name = f"짱베이스볼_야구바지_재고발주_v5_20260319_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    shutil.copy2(EXCEL_PATH, BACKUP_DIR / backup_name)
    log.info(f"백업 완료: {backup_name}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["재고현황_발주"]

    # 타임스탬프 업데이트
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(row=2, column=1).value = f"{timestamp} | 고도몰 Open API 실시간 | 26개 상품 | 초록=주문중 입력 | 골드=발주 필요"

    # 전체 집계 변수
    total_stock = 0
    total_order = 0
    oos_options = []     # 품절 옵션
    low_options = []     # 부족 옵션 (5개 이하)
    reorder_products = 0 # 발주 필요 상품 수
    reorder_options = 0  # 발주 필요 옵션 수
    all_products = []

    for row in range(6, ws.max_row + 1):
        gno = ws.cell(row=row, column=2).value
        name = ws.cell(row=row, column=3).value

        if not gno or not str(gno).isdigit():
            continue

        gno = str(gno)

        # API 데이터가 있으면 현재 재고 업데이트
        if gno in api_data:
            api_sizes = api_data[gno]["sizes"]

            # 사이즈별 재고 업데이트 (C5~C11)
            row_total = 0
            for i, size in enumerate(SIZES):
                # 이 상품의 옵션에서 해당 사이즈 찾기
                stock = 0
                for opt_key, opt_val in api_sizes.items():
                    # 옵션값에서 사이즈 번호 추출 (예: "28", "30 (누빔없음)" 등)
                    if opt_key.startswith(size) or opt_key == size:
                        stock = opt_val
                        break
                    # 정확한 매칭이 안 되면 사이즈 번호가 포함된 경우
                    if size in opt_key.split("(")[0].strip().split(" ")[0]:
                        stock = opt_val

                ws.cell(row=row, column=STOCK_COL_START + i).value = stock
                row_total += stock

            # 합계 업데이트
            ws.cell(row=row, column=STOCK_TOTAL_COL).value = row_total

        # 현재 값 읽기 (업데이트 후)
        product_data = {"goodsNo": gno, "name": name, "sizes": {}, "order": {}, "need": {}, "reorder": {}}
        prod_stock_total = 0
        prod_order_total = 0
        prod_has_reorder = False

        for i, size in enumerate(SIZES):
            stock = ws.cell(row=row, column=STOCK_COL_START + i).value or 0
            order = ws.cell(row=row, column=ORDER_COL_START + i).value or 0
            need = ws.cell(row=row, column=NEED_COL_START + i).value or 0

            # 추가 발주 계산: MAX(필요 - 현재 - 주문중, 0)
            reorder = max(need - stock - order, 0)
            ws.cell(row=row, column=REORDER_COL_START + i).value = reorder

            product_data["sizes"][size] = stock
            product_data["order"][size] = order
            product_data["need"][size] = need
            product_data["reorder"][size] = reorder

            prod_stock_total += stock
            prod_order_total += order

            if stock == 0:
                oos_options.append({"name": name, "size": size})
            elif stock <= 5:
                low_options.append({"name": name, "size": size, "qty": stock})

            if reorder > 0:
                reorder_options += 1
                prod_has_reorder = True

        # 합계 열 업데이트
        ws.cell(row=row, column=ORDER_TOTAL_COL).value = prod_order_total
        reorder_sum = sum(product_data["reorder"].values())
        ws.cell(row=row, column=REORDER_TOTAL_COL).value = reorder_sum

        total_stock += prod_stock_total
        total_order += prod_order_total
        if prod_has_reorder:
            reorder_products += 1

        product_data["stock_total"] = prod_stock_total
        product_data["order_total"] = prod_order_total
        product_data["reorder_total"] = reorder_sum
        all_products.append(product_data)

    # 엑셀 저장
    wb.save(EXCEL_PATH)
    log.info(f"엑셀 업데이트 완료: {EXCEL_PATH}")

    # 결과용 엑셀도 output에 복사
    output_excel = OUTPUT_DIR / "stock-check-result.xlsx"
    shutil.copy2(EXCEL_PATH, output_excel)

    return {
        "timestamp": timestamp,
        "total_products": len(all_products),
        "total_stock": total_stock,
        "total_order": total_order,
        "oos_options": oos_options,
        "low_options": low_options,
        "oos_count": len(oos_options),
        "low_count": len(low_options),
        "reorder_products": reorder_products,
        "reorder_options": reorder_options,
        "all_products": all_products
    }


def build_html_report(analysis):
    """기존 스크린샷과 동일한 형태의 HTML 리포트 생성"""
    ts = analysis["timestamp"]
    total = analysis["total_stock"]
    order = analysis["total_order"]
    oos = analysis["oos_count"]
    low = analysis["low_count"]
    rp = analysis["reorder_products"]
    ro = analysis["reorder_options"]
    prods = analysis["total_products"]

    # 품절 뱃지
    oos_badges = ""
    for item in analysis["oos_options"]:
        oos_badges += (
            f'<span style="display:inline-block;padding:6px 14px;margin:4px;'
            f'border:1px solid #ddd;border-radius:20px;font-size:13px;'
            f'color:#555;">{item["name"]} {item["size"]}</span>\n'
        )

    # 부족 뱃지 (상위 15건)
    low_sorted = sorted(analysis["low_options"], key=lambda x: x["qty"])
    low_badges = ""
    for item in low_sorted[:15]:
        low_badges += (
            f'<span style="display:inline-block;padding:6px 14px;margin:4px;'
            f'border:1px solid #ddd;border-radius:20px;font-size:13px;'
            f'color:#555;">{item["name"]} {item["size"]} ({item["qty"]})</span>\n'
        )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head><body style="margin:0;padding:0;background:#f5f5f5;">
<div style="max-width:700px;margin:20px auto;background:#fff;font-family:'Noto Sans KR',-apple-system,sans-serif;">

  <div style="padding:28px 32px;border-left:4px solid #C62828;">
    <div style="font-size:22px;font-weight:800;color:#1B2A4A;">짱베이스볼 야구바지 재고 현황</div>
    <div style="font-size:13px;color:#888;margin-top:6px;">{ts} | 고도몰 Open API 실시간 | {prods}개 상품</div>
  </div>

  <table style="width:100%;border-collapse:collapse;margin:0;">
    <tr>
      <td style="padding:16px;text-align:center;background:#f8f9fa;border:1px solid #eee;">
        <div style="font-size:11px;color:#888;font-weight:600;">전체 재고</div>
        <div style="font-size:22px;font-weight:800;color:#1B2A4A;margin-top:4px;">{total:,}개</div>
      </td>
      <td style="padding:16px;text-align:center;background:#f8f9fa;border:1px solid #eee;">
        <div style="font-size:11px;color:#888;font-weight:600;">주문중</div>
        <div style="font-size:22px;font-weight:800;color:#1B2A4A;margin-top:4px;">{order:,}개</div>
      </td>
      <td style="padding:16px;text-align:center;background:#f8f9fa;border:1px solid #eee;">
        <div style="font-size:11px;color:#888;font-weight:600;">품절</div>
        <div style="font-size:22px;font-weight:800;color:#C62828;margin-top:4px;">{oos}건</div>
      </td>
      <td style="padding:16px;text-align:center;background:#f8f9fa;border:1px solid #eee;">
        <div style="font-size:11px;color:#888;font-weight:600;">부족</div>
        <div style="font-size:22px;font-weight:800;color:#E65100;margin-top:4px;">{low}건</div>
      </td>
      <td style="padding:16px;text-align:center;background:#D4A843;border:1px solid #C09838;">
        <div style="font-size:11px;color:#1B2A4A;font-weight:600;">발주 필요</div>
        <div style="font-size:22px;font-weight:800;color:#1B2A4A;margin-top:4px;">{rp}개 상품 / {ro}개</div>
      </td>
    </tr>
  </table>

  {"" if oos == 0 else f'''
  <div style="padding:24px 32px;">
    <div style="font-size:16px;font-weight:700;color:#C62828;margin-bottom:12px;">품절 사이즈 ({oos}건)</div>
    <div>{oos_badges}</div>
  </div>'''}

  {"" if low == 0 else f'''
  <div style="padding:0 32px 24px;">
    <div style="font-size:16px;font-weight:700;color:#E65100;margin-bottom:12px;">부족 사이즈 ({low}건)</div>
    <div>{low_badges}</div>
  </div>'''}

  <div style="padding:16px 32px;background:#f8f9fa;font-size:12px;color:#999;border-top:1px solid #eee;">
    짱베이스볼 재고 자동 모니터링 (v8 Open API) | 매일 오전 8시 자동 실행
  </div>

</div>
</body></html>"""

    return html


def send_gmail(config, subject, html_body):
    """Gmail SMTP로 메일 발송 (엑셀 첨부)"""
    email_config = config.get("email", {})
    gmail_addr = email_config.get("gmail_address", "")
    app_password = email_config.get("gmail_app_password", "")

    if not gmail_addr or not app_password:
        log.warning("Gmail 설정 없음 - 메일 발송 건너뜀")
        return False

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = gmail_addr
    msg["To"] = gmail_addr

    # HTML 본문
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # 엑셀 첨부
    excel_path = OUTPUT_DIR / "stock-check-result.xlsx"
    if excel_path.exists():
        with open(excel_path, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            filename = f"재고현황_{datetime.now().strftime('%Y%m%d')}.xlsx"
            part.add_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
            msg.attach(part)
            log.info(f"엑셀 첨부: {filename}")

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as server:
            server.login(gmail_addr, app_password)
            server.send_message(msg)
        log.info(f"메일 발송 완료: {gmail_addr}")
        return True
    except smtplib.SMTPAuthenticationError:
        log.error("Gmail 인증 실패 - 앱 비밀번호를 확인하세요")
        return False
    except Exception as e:
        log.error(f"메일 발송 실패: {e}")
        return False


def save_json(analysis):
    """결과 JSON 저장"""
    output = {
        "timestamp": analysis["timestamp"],
        "version": "v8",
        "method": "godomall_openhub_api",
        "summary": {
            "total_products": analysis["total_products"],
            "total_stock": analysis["total_stock"],
            "total_order": analysis["total_order"],
            "out_of_stock_options": analysis["oos_count"],
            "low_stock_options": analysis["low_count"],
            "reorder_products": analysis["reorder_products"],
            "reorder_options": analysis["reorder_options"]
        },
        "all_products": analysis["all_products"]
    }
    output_path = OUTPUT_DIR / "stock-check-result.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    log.info(f"JSON 저장: {output_path}")


def main():
    args = sys.argv[1:]
    no_mail = "--nomail" in args
    test_mode = "--test" in args

    log.info("=" * 50)
    log.info("짱베이스볼 재고 모니터링 시작")
    log.info("=" * 50)

    # 1. 설정 로드
    config = load_config()
    log.info(f"모니터링 대상: {len(config['monitor_products'])}개 상품")

    if test_mode:
        config["monitor_products"] = config["monitor_products"][:1]
        log.info("[테스트 모드] 첫 번째 상품만 조회")

    # 2. API로 재고 조회
    api_data = fetch_stock_api(config)
    if not api_data:
        log.error("API 조회 결과 없음 - 종료")
        sys.exit(1)
    log.info(f"API 조회 완료: {len(api_data)}개 상품")

    # 3. 엑셀 업데이트 + 분석
    analysis = update_excel(api_data)
    if not analysis:
        log.error("엑셀 업데이트 실패 - 종료")
        sys.exit(1)

    log.info(f"전체 재고: {analysis['total_stock']:,}개 | 주문중: {analysis['total_order']:,}개")
    log.info(f"품절: {analysis['oos_count']}건 / 부족: {analysis['low_count']}건")
    log.info(f"발주 필요: {analysis['reorder_products']}개 상품 / {analysis['reorder_options']}개 옵션")

    # 4. JSON 저장
    save_json(analysis)

    # 5. HTML 리포트 생성
    html = build_html_report(analysis)
    html_path = OUTPUT_DIR / "stock-check-report.html"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    log.info(f"HTML 리포트 저장: {html_path}")

    # 6. 메일 발송
    if not no_mail:
        oos = analysis["oos_count"]
        low = analysis["low_count"]
        prefix = config["email"].get("subject_prefix", "[짱베이스볼 재고알림]")
        today = datetime.now().strftime("%Y-%m-%d")
        subject = f"{prefix} {today} 재고 현황 (품절 {oos}건 / 부족 {low}건)"
        sent = send_gmail(config, subject, html)
        if not sent:
            log.warning("메일 미발송 - HTML 파일로 확인하세요")
    else:
        log.info("[--nomail] 메일 발송 건너뜀")

    # 7. 완료
    log.info("재고 모니터링 완료")
    print(f"\n전체 재고: {analysis['total_stock']:,}개 | 주문중: {analysis['total_order']:,}개 | 품절: {analysis['oos_count']}건 | 부족: {analysis['low_count']}건 | 발주: {analysis['reorder_products']}개 상품")


if __name__ == "__main__":
    main()
